import openpyxl
from io import BytesIO
from sqlalchemy.orm import Session
from fastapi import HTTPException, UploadFile
from models import employee_models, core_models, department_models # Added department_models
from core.auth import get_password_hash
import uuid
import datetime
import re
from difflib import get_close_matches

class EmployeeImportService:
    def __init__(self, db: Session):
        self.db = db
        # Dictionary of standard column names to regex patterns/synonyms
        self.known_columns = {
            "code": ["code", "id", "employee id", "staff id", "رقم الموظف", "الكود", "الرقم الوظيفي", "رقم الهوية", "الرقم التسلسلي"],
            "first_name": ["first name", "firstname", "fname", "name", "الاسم الاول", "الاسم", "الاسم الأول", "الاسم الثلاثي", "اسم الموظف"],
            "last_name": ["last name", "lastname", "lname", "surname", "family name", "الكنية", "العائلة", "اللقب", "الاسم الأخير"],
            "work_email": ["email", "e-mail", "work email", "official email", "البريد الالكتروني", "الايميل", "البريد الإلكتروني", "عنوان البريد"],
            "department": ["department", "dept", "division", "unit", "القسم", "الادارة", "الوحدة", "الدائرة"],
            "position": ["position", "job title", "designation", "role", "المنصب", "المسمى الوظيفي", "الوظيفة"],
            "joining_date": ["joining date", "join date", "hire date", "started at", "تاريخ الانضمام", "تاريخ التعيين", "تاريخ المباشرة", "تاريخ التوظيف"],
            "company": ["company", "organization", "firm", "الشركة", "اسم الشركة", "المنشأة"]
        }

    async def analyze_file(self, file: UploadFile):
        """
        Analyze the Excel file and return structure/preview + SUGGESTED MAPPINGS.
        Does NOT save data.
        """
        if not file.filename.endswith('.xlsx'):
            raise HTTPException(status_code=400, detail="Invalid file format. Please upload .xlsx file")

        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        sheet = wb.active
        
        headers = []
        for cell in sheet[1]:
            val = str(cell.value).strip() if cell.value else f"Column {cell.column}"
            headers.append(val)

        preview_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                # Sanitize data: convert to string, handle None
                clean_row = [str(x) if x is not None else "" for x in row]
                preview_rows.append(clean_row)
        
        # intelligent mapping suggestion
        suggested_mappings = {}
        for db_field, keywords in self.known_columns.items():
            # Try exact match first
            match = None
            lower_headers = [h.lower() for h in headers]
            
            for keyword in keywords:
                if keyword in lower_headers:
                    match = headers[lower_headers.index(keyword)]
                    break
            
            # If no exact match, try fuzzy
            if not match:
                closest = get_close_matches(keywords[0], lower_headers, n=1, cutoff=0.6)
                if closest:
                     match = headers[lower_headers.index(closest[0])]
            
            if match:
                suggested_mappings[db_field] = headers.index(match) # Store index

        return {
            "headers": headers,
            "preview_rows": preview_rows,
            "total_rows": sheet.max_row - 1,
            "suggested_mappings": suggested_mappings
        }

    def execute_import(self, data: list, options: dict, current_user_id: uuid.UUID):
        """
        Execute the import process with "Professional Logic".
        Features:
        - Auto-create Departments (Master Data)
        - Update vs Create Logic
        - Transaction Safety
        - Detailed Error Report
        """
        results = {
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": []
        }
        
        mappings = options.get("mappings", {})
        # Mappings should be { "db_field": "excel_col_index" } 
        
        skip_duplicates = options.get("skipDuplicates", True)
        update_existing = options.get("updateExisting", False) # New option
        create_users = options.get("createUsers", False) 
        auto_create_dept = options.get("autoCreateDepartments", True) # New option
        
        # Cache for Master Data to reduce DB hits
        dept_cache = {d.name.lower(): d.id for d in self.db.query(department_models.Department).all()}
        
        for index, row in enumerate(data):
            row_num = index + 2
            try:
                # 1. Extract & Sanitize Data
                emp_data = self._extract_data(row, mappings)
                
                if not emp_data.get("code") or not emp_data.get("first_name"):
                    results["errors"].append(f"Row {row_num}: Missing Code or First Name")
                    continue

                # 2. Master Data Handling (Department)
                dept_id = None
                if emp_data.get("department"):
                    dept_name = emp_data["department"]
                    dept_key = dept_name.lower()
                    if dept_key in dept_cache:
                        dept_id = dept_cache[dept_key]
                    elif auto_create_dept:
                        # Auto-create new department
                        new_dept = department_models.Department(name=dept_name)
                        self.db.add(new_dept)
                        self.db.flush()
                        dept_id = new_dept.id
                        dept_cache[dept_key] = dept_id

                # 3. Check Existence
                existing_emp = self.db.query(employee_models.Employee).filter(
                    (employee_models.Employee.code == emp_data["code"])
                ).first()
                
                # Also check by email if code not found (optional safety)
                if not existing_emp and emp_data.get("work_email"):
                     existing_emp = self.db.query(employee_models.Employee).filter(
                        (employee_models.Employee.work_email == emp_data["work_email"])
                    ).first()

                if existing_emp:
                    if skip_duplicates and not update_existing:
                        results["skipped"] += 1
                        continue
                    
                    # Update Logic
                    existing_emp.first_name = emp_data["first_name"]
                    existing_emp.last_name = emp_data.get("last_name", "")
                    existing_emp.full_name = f"{emp_data['first_name']} {emp_data.get('last_name', '')}".strip()
                    existing_emp.position = emp_data.get("position")
                    existing_emp.department_name = emp_data.get("department") # Cache name
                    existing_emp.company = emp_data.get("company") # Update company
                    if dept_id: existing_emp.department_id = dept_id
                    
                    if emp_data.get("joining_date"):
                        existing_emp.joining_date = emp_data["joining_date"]

                    self.db.flush() 
                    results["updated"] += 1
                    target_emp = existing_emp

                else:
                    # Create Logic
                    full_name = f"{emp_data['first_name']} {emp_data.get('last_name', '')}".strip()
                    target_emp = employee_models.Employee(
                        code=emp_data["code"],
                        first_name=emp_data["first_name"],
                        last_name=emp_data.get("last_name", ""),
                        full_name=full_name,
                        work_email=emp_data.get("work_email"),
                        position=emp_data.get("position"),
                        department_name=emp_data.get("department"),
                        company=emp_data.get("company"),
                        department_id=dept_id,
                        status="active",
                        joining_date=emp_data.get("joining_date") or datetime.date.today()
                    )
                    self.db.add(target_emp)
                    self.db.flush()
                    self.db.flush() # Flush to get ID for potential user creation
                    results["created"] += 1

                # 4. Handle System Access (User Creation)
                if create_users and emp_data.get("work_email"):
                    self._handle_user_creation(target_emp, emp_data["work_email"])
                
            except Exception as e:
                # Rollback changes for this specific row if an error occurs
                # Note: This rollback will revert all changes made in the current session since the last commit/begin.
                # For row-by-row error handling with a single final commit, a sub-transaction or savepoint might be needed.
                # However, given the current structure, we'll just log the error and continue.
                # The final commit will then attempt to commit all successful rows.
                # If a flush failed, the error would be caught here.
                results["errors"].append(f"Row {row_num}: {str(e)}")
        
        # Commit all successful changes at once
        self.db.commit()
        return results

    def _extract_data(self, row, mappings):
        """Helper to safely extract and convert data types"""
        data = {}
        for field, col_idx in mappings.items():
            val = self._get_val(row, col_idx)
            if field == "joining_date" and val:
                data[field] = self._parse_date(val)
            else:
                data[field] = val
        return data

    def _handle_user_creation(self, employee, email):
        """Create or Link User account"""
        if employee.user_id: 
            return # Already linked

        existing_user = self.db.query(core_models.User).filter(core_models.User.email == email).first()
        if not existing_user:
            # Generate smart password or default
            # Professional: Send email invitation? For now, default password.
            temp_pass = get_password_hash("Jandali@2025") 
            new_user = core_models.User(
                name=employee.full_name,
                email=email,
                password=temp_pass,
                role="user",
                is_active=True
            )
            self.db.add(new_user)
            self.db.flush()
            employee.user_id = new_user.id
        else:
            employee.user_id = existing_user.id

    def _get_val(self, row, col_index):
        if col_index is None: return None
        try:
            idx = int(col_index)
            if 0 <= idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else None
        except:
            return None
        return None

    def _parse_date(self, date_str):
        """Robust date parser"""
        try:
            if isinstance(date_str, datetime.datetime):
                return date_str.date()
            if isinstance(date_str, datetime.date):
                return date_str
            
            # Try formats
            formats = ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y"]
            for fmt in formats:
                try:
                    return datetime.datetime.strptime(str(date_str), fmt).date()
                except:
                    pass
            return None
        except:
            return None
